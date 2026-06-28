import pandas as pd
import re
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from models_document import ParsedDocument
from logger import logger

# Styling Constants
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")
_BORDER_SIDE = Side(style="thin", color="BFBFBF")
_CELL_BORDER = Border(left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE)
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

def export_to_excel(doc: ParsedDocument, output_path: str) -> None:
    wb = Workbook()
    
    # 1. Summary Sheet
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Property", "Value"])
    summary.append(["Document Type", doc.doc_type.upper()])
    summary.append(["Total Pages", doc.metadata.get("pages", "Unknown")])
    summary.append(["Extractor Used", doc.metadata.get("method", "Auto-Classifier")])
    _style_header_row(summary, 1, 2)

    # 2. Tables Sheets (Each large table gets its own sheet)
    used_sheet_names = {"Summary", "Key Values", "Sections", "Products", "Raw Text"}
    if doc.tables:
        for i, table in enumerate(doc.tables, 1):
            title = table.get("title", f"Table_{i}")
            sheet_name = _make_sheet_name(title, i, used_sheet_names)
            used_sheet_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)
            df = table["dataframe"]
            _write_df_to_sheet(ws, df)

    # 3. Key Values Sheet
    if doc.key_values:
        ws = wb.create_sheet("Key Values")
        df = pd.DataFrame(doc.key_values)
        _write_df_to_sheet(ws, df)

    # 4. Sections Sheet
    if doc.sections:
        ws = wb.create_sheet("Sections")
        df = pd.DataFrame(doc.sections)
        _write_df_to_sheet(ws, df)

    # 5. Products Sheet
    if doc.products:
        ws = wb.create_sheet("Products")
        df = pd.DataFrame(doc.products)
        _write_df_to_sheet(ws, df)

    # 6. Raw Text Sheet
    if doc.text:
        ws = wb.create_sheet("Raw Text")
        # Split text to avoid Excel cell limit
        lines = doc.text.splitlines()
        df = pd.DataFrame({"Content": lines})
        _write_df_to_sheet(ws, df)

    # Global Formatting
    for sheet in wb.sheetnames:
        _autofit_columns(wb[sheet])
        wb[sheet].freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Universal Excel exported to {output_path}")

def _write_df_to_sheet(ws, df: pd.DataFrame):
    headers = df.columns.tolist()
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for _, row in df.iterrows():
        ws.append([_try_numeric(str(v)) for v in row])

def _style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER

def _autofit_columns(ws, max_width: int = 60):
    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_len + 2, max_width)

def _try_numeric(value: str) -> Any:
    v = value.strip().replace(",", "")
    try:
        if "." in v: return float(v)
        return int(v)
    except: return value

def _make_sheet_name(title: str, idx: int, used: set) -> str:
    name = re.sub(r"[\\/:*?\[\]]", "", str(title))[:28].strip() or f"Sheet_{idx}"
    while name in used:
        name = name[:25] + f"_{idx}"
        idx += 1
    return name
